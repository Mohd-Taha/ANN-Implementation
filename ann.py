import math
import random
import xlsxwriter
from openpyxl import load_workbook
from threading import Thread


class ThreadWithReturnValue(Thread):
    def __init__(self, group=None, target=None, name=None, args=(), kwargs={}, Verbose=None):
        Thread.__init__(self, group, target, name, args, kwargs)
        self._return = None

    def run(self):
        # print(type(self._target))
        if self._target is not None:
            self._return = self._target(*self._args, **self._kwargs)

    def join(self, *args):
        Thread.join(self, *args)
        return self._return


def dataReading(path):
    # loading workbook, then working on active sheet
    dataFile = load_workbook(path)
    dataSheet = dataFile.active

    # then getting its rows (1st row -> headers), then reading other (generator object)
    dataRows = dataSheet.rows
    dataHeaders = [cell.value for cell in next(dataRows)]
    # print(dataHeaders)
    dataSet = []
    for row in dataRows:
        record = []
        for cell in row:
            record.append(float(cell.value) / 15000)
        dataSet.append(record)
        # print(record)
    return dataSet


def dataOutput(i, result):
    outputSheet.write('A' + str(i), result[0])
    outputSheet.write('B' + str(i), result[1])
    outputSheet.write('C' + str(i), result[2])
    outputSheet.write('D' + str(i), result[3])
    outputSheet.write('E' + str(i), result[4])
    outputSheet.write('F' + str(i), result[5])
    outputSheet.write('G' + str(i), result[6])
    outputSheet.write('H' + str(i), result[7])
    outputSheet.write('I' + str(i), result[8])
    outputSheet.write('J' + str(i), result[9])
    outputSheet.write('K' + str(i), result[10])
    outputSheet.write('L' + str(i), result[11])
    outputSheet.write('M' + str(i), result[12])
    outputSheet.write('N' + str(i), result[13])
    outputSheet.write('O' + str(i), result[14])
    outputSheet.write('P' + str(i), result[15])


def calculateOutput(i, ix_1, ix_2, ix_3, iw_14, iw_15, iw_24, iw_25, iw_34, iw_35, iw_46, iw_56, ib_4, ib_5, ib_6, iy):
    thresholdError = 0.0001
    iErr_6 = 1
    j = 1
    while (abs(iErr_6) > thresholdError):

        # calculating outputs of intermediate neurons
        iI_4 = ix_1 * iw_14 + ix_2 * iw_24 + ix_3 * iw_34 + ib_4
        iI_5 = ix_1 * iw_15 + ix_2 * iw_25 + ix_3 * iw_35 + ib_5
        iO_4 = 1 / (1 + math.e**(-iI_4))
        iO_5 = 1 / (1 + math.e**(-iI_5))

        # calculating output
        iI_6 = iO_4 * iw_46 + iO_5 * iw_56 + ib_6
        iO_6 = 1 / (1 + math.e**(-iI_6))

        # calculating output error => j_err = j_O * (1 - j_O) * (iy - j_O)
        iErr_6 = iO_6 * (1-iO_6) * (iy-iO_6)
        # print("Input No#: ", str(i), ",", "\tIteration No#: ",
        #       str(j), ",", "\tOutput Error: ", str(iErr_6))

        if (abs(iErr_6) > thresholdError):

            # calculating hidden error => j_err = j_O * (1 - j_O) * (k_err * jk_weight)
            iErr_4 = iO_4 * (1 - iO_4) * iw_46 * iErr_6
            iErr_5 = iO_5 * (1 - iO_5) * iw_56 * iErr_6

            # adjusting biases => j_bias = j_bias + rate * j_error
            ib_6 += learningRate * iErr_6
            ib_5 += learningRate * iErr_5
            ib_4 += learningRate * iErr_4

            # adjusting weights => ij_weight = ij_weight + rate * j_error * i_input
            iw_56 += learningRate * iErr_6 * iO_5
            iw_46 += learningRate * iErr_6 * iO_4
            iw_35 += learningRate * iErr_5 * ix_3
            iw_34 += learningRate * iErr_4 * ix_3
            iw_25 += learningRate * iErr_5 * ix_2
            iw_24 += learningRate * iErr_4 * ix_2
            iw_15 += learningRate * iErr_5 * ix_1
            iw_14 += learningRate * iErr_4 * ix_1

        j += 1

    return [ix_1, ix_2, ix_3, iw_14, iw_24, iw_34, ib_4, iw_15, iw_25, iw_35, ib_5, iw_46, iw_56, ib_6, iO_6, iErr_6]


if __name__ == "__main__":
    pathToFile = input("Paste Path to File (with double backslash and file name)\t:")
    dataSet = dataReading(pathToFile)
    w_14 = random.uniform(-1.0, 1.0)
    w_15 = random.uniform(-1.0, 1.0)
    w_24 = random.uniform(-1.0, 1.0)
    w_25 = random.uniform(-1.0, 1.0)
    w_34 = random.uniform(-1.0, 1.0)
    w_35 = random.uniform(-1.0, 1.0)
    w_46 = random.uniform(-1.0, 1.0)
    w_56 = random.uniform(-1.0, 1.0)
    b_4 = random.uniform(-1.0, 1.0)
    b_5 = random.uniform(-1.0, 1.0)
    b_6 = random.uniform(-1.0, 1.0)
    x_1 = 0
    x_2 = 0
    x_3 = 0
    y = 0
    learningRate = 1
    epoch = 1
    previousAverage = [w_14, w_24, w_34, b_4,
                       w_15, w_25, w_35, b_5, w_46, w_56, b_6]

    while (True):
        fileName = 'annIntermediate' + str(epoch) + '.xlsx'
        outputFile = xlsxwriter.Workbook(fileName)
        threads = []

        print("\tEPOCH:\t"+str(epoch)+'\n')

        # creating threads with their parameters
        for i, data in enumerate(dataSet):
            x_1, x_2, x_3, y = data
            t = ThreadWithReturnValue(target=calculateOutput, args=(
                i+1, x_1, x_2, x_3, w_14, w_15, w_24, w_25, w_34, w_35, w_46, w_56, b_4, b_5, b_6, y))
            threads.append(t)

        w_14 = w_15 = w_24 = w_25 = w_34 = w_35 = w_46 = w_56 = b_4 = b_5 = b_6 = 0

        sheetname = 'annIntermediate' + str(epoch)
        outputSheet = outputFile.add_worksheet(sheetname)
        outputSheet.write('A1', 'Inp_X1')
        outputSheet.write('B1', 'Inp_X2')
        outputSheet.write('C1', 'Inp_X3')
        outputSheet.write('D1', 'W_14')
        outputSheet.write('E1', 'W_24')
        outputSheet.write('F1', 'W_34')
        outputSheet.write('G1', 'B_4')
        outputSheet.write('H1', 'W_15')
        outputSheet.write('I1', 'W_25')
        outputSheet.write('J1', 'W_35')
        outputSheet.write('K1', 'B_5')
        outputSheet.write('L1', 'W_46')
        outputSheet.write('M1', 'W_56')
        outputSheet.write('N1', 'B_6')
        outputSheet.write('O1', 'Calc_Y')

        # threads started
        for t in threads:
            t.start()

        # waiting for threads to complete and sum their results
        for i, t in enumerate(threads):
            res = t.join()
            w_14 += res[3]
            w_24 += res[4]
            w_34 += res[5]
            b_4 += res[6]
            w_15 += res[7]
            w_25 += res[8]
            w_35 += res[9]
            b_5 += res[10]
            w_46 += res[11]
            w_56 += res[12]
            b_6 += res[13]
            dataOutput(i+2, res)

        # average of all weights and biases
        n = len(threads)
        w_14 /= n
        w_24 /= n
        w_34 /= n
        b_4 /= n
        w_15 /= n
        w_25 /= n
        w_35 /= n
        b_5 /= n
        w_46 /= n
        w_56 /= n
        b_6 /= n

        currentAverage = [w_14, w_24, w_34, b_4,
                          w_15, w_25, w_35, b_5, w_46, w_56, b_6]

        endTask = True
        # all inputs gave absolute error less than threshold in first go, results are here !!
        for i in range(len(currentAverage)):
            if currentAverage[i] != previousAverage[i]:
                endTask = False
                previousAverage[i] = currentAverage[i]
                break

        if (endTask):
            resultFile = xlsxwriter.Workbook('annResult.xlsx')
            resultSheet = resultFile.add_worksheet('annResult')
            resultSheet.write('A1', 'W_14')
            resultSheet.write('B1', 'W_24')
            resultSheet.write('C1', 'W_34')
            resultSheet.write('D1', 'B_4')
            resultSheet.write('E1', 'W_15')
            resultSheet.write('F1', 'W_25')
            resultSheet.write('G1', 'W_35')
            resultSheet.write('H1', 'B_5')
            resultSheet.write('I1', 'W_46')
            resultSheet.write('J1', 'W_56')
            resultSheet.write('K1', 'B_6')
            resultSheet.write('A2', w_14)
            resultSheet.write('B2', w_24)
            resultSheet.write('C2', w_34)
            resultSheet.write('D2', b_4)
            resultSheet.write('E2', w_15)
            resultSheet.write('F2', w_25)
            resultSheet.write('G2', w_35)
            resultSheet.write('H2', b_5)
            resultSheet.write('I2', w_46)
            resultSheet.write('J2', w_56)
            resultSheet.write('K2', b_6)
            resultFile.close()
            break

        epoch += 1
        learningRate *= 0.95
        if (learningRate < 0.4):
            learningRate = 0.4
        outputFile.close()
